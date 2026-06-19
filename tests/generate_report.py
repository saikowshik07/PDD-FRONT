import os
import sys
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "results", "results.json")
    excel_path = os.path.join(script_dir, "results", "E2E_Test_Report.xlsx")
    
    if not os.path.exists(json_path):
        print(f"Error: results.json not found at {json_path}")
        sys.exit(1)
        
    with open(json_path, "r", encoding="utf-8") as f:
        results = json.load(f)
        
    wb = Workbook()
    
    # ----------------------------------------------------
    # Styles Definition
    # ----------------------------------------------------
    primary_color = "1F4E79"  # Deep Blue
    accent_color = "D9E1F2"   # Soft Blue-Gray
    text_color = "FFFFFF"     # White
    pass_color = "C6EFCE"     # Soft Green
    pass_text = "006100"
    fail_color = "FFC7CE"     # Soft Red
    fail_text = "9C0006"
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color=text_color)
    header_font = Font(name="Segoe UI", size=11, bold=True, color=primary_color)
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)
    
    title_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # ----------------------------------------------------
    # TAB 1: EXECUTIVE SUMMARY
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Header Banner
    ws_summary.merge_cells("A1:D2")
    cell_a1 = ws_summary["A1"]
    cell_a1.value = "SignVision AI — Quality Assurance Executive Dashboard"
    cell_a1.font = title_font
    cell_a1.fill = title_fill
    cell_a1.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata Block
    metadata = [
        ("Execution Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Test Environment", "GitHub Actions (CI/CD Pipeline)"),
        ("Source Branch", os.environ.get("GITHUB_REF_NAME", "Local / Main")),
        ("Commit SHA", os.environ.get("GITHUB_SHA", "Local-Run")[:8]),
    ]
    
    ws_summary.cell(row=4, column=1, value="METADATA").font = header_font
    for idx, (k, v) in enumerate(metadata):
        row = 5 + idx
        ws_summary.cell(row=row, column=1, value=k).font = bold_font
        ws_summary.cell(row=row, column=1).border = thin_border
        
        ws_summary.cell(row=row, column=2, value=v).font = regular_font
        ws_summary.cell(row=row, column=2).border = thin_border
        
    # Test Statistics
    total_tests = len(results)
    passed_tests = sum(1 for r in results if r["status"] == "PASS")
    failed_tests = total_tests - passed_tests
    pass_rate = (passed_tests / total_tests) if total_tests > 0 else 0
    
    # Count per sheet category
    functional_count = sum(1 for r in results if r["category"] == "Functionality")
    vuln_count = sum(1 for r in results if r["category"] == "Security / Vulnerability")
    api_unit_count = sum(1 for r in results if r["category"] == "API Unit")
    ui_ux_count = sum(1 for r in results if r["category"] == "UI UX")
    
    ws_summary.cell(row=4, column=3, value="RESULTS SUMMARY").font = header_font
    stats = [
        ("Total Test Cases", total_tests),
        ("Passed Cases", passed_tests),
        ("Failed Cases", failed_tests),
        ("Overall Pass Rate", f"{pass_rate:.1%}")
    ]
    
    for idx, (k, v) in enumerate(stats):
        row = 5 + idx
        c1 = ws_summary.cell(row=row, column=3, value=k)
        c1.font = bold_font
        c1.border = thin_border
        
        c2 = ws_summary.cell(row=row, column=4, value=v)
        c2.font = bold_font if k == "Overall Pass Rate" else regular_font
        c2.border = thin_border
        
        if k == "Overall Pass Rate":
            if pass_rate == 1.0:
                c2.fill = PatternFill(start_color=pass_color, end_color=pass_color, fill_type="solid")
                c2.font = Font(name="Segoe UI", size=11, bold=True, color=pass_text)
            else:
                c2.fill = PatternFill(start_color=fail_color, end_color=fail_color, fill_type="solid")
                c2.font = Font(name="Segoe UI", size=11, bold=True, color=fail_text)
                
    # Distribution matrix
    start_row = 11
    ws_summary.cell(row=start_row, column=1, value="TEST TYPE DISTRIBUTION").font = header_font
    dist_stats = [
        ("Functional Tests", functional_count),
        ("Vulnerability Tests", vuln_count),
        ("API Unit Tests", api_unit_count),
        ("UI UX E2E Tests", ui_ux_count),
    ]
    for idx, (name, count) in enumerate(dist_stats):
        row = start_row + 1 + idx
        ws_summary.cell(row=row, column=1, value=name).font = bold_font
        ws_summary.cell(row=row, column=1).border = thin_border
        ws_summary.cell(row=row, column=2, value=count).font = regular_font
        ws_summary.cell(row=row, column=2).border = thin_border

    # Helper function to create worksheets for each divider
    def create_categorized_sheet(title, filter_category=None):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        
        headers = ["No.", "Category", "Test Case Name", "Description", "Status", "Duration (s)", "Execution Details"]
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
        
        # Filter tests
        if filter_category:
            filtered_results = [r for r in results if r["category"] == filter_category]
        else:
            filtered_results = results  # All results
            
        for idx, r in enumerate(filtered_results, 1):
            row = 1 + idx
            ws.row_dimensions[row].height = 22
            
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=r["category"])
            ws.cell(row=row, column=3, value=r["name"]).font = bold_font
            ws.cell(row=row, column=4, value=r["description"])
            
            status_cell = ws.cell(row=row, column=5, value=r["status"])
            status_cell.alignment = Alignment(horizontal="center")
            if r["status"] == "PASS":
                status_cell.fill = PatternFill(start_color=pass_color, end_color=pass_color, fill_type="solid")
                status_cell.font = Font(name="Segoe UI", size=11, bold=True, color=pass_text)
            else:
                status_cell.fill = PatternFill(start_color=fail_color, end_color=fail_color, fill_type="solid")
                status_cell.font = Font(name="Segoe UI", size=11, bold=True, color=fail_text)
                
            ws.cell(row=row, column=6, value=r["duration"]).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=7, value=r["details"])
            
            for c in range(1, 8):
                cell = ws.cell(row=row, column=c)
                cell.border = thin_border
                if c != 3 and c != 5:
                    cell.font = regular_font

    # ----------------------------------------------------
    # TAB 2: FUNCTIONAL
    # ----------------------------------------------------
    create_categorized_sheet("Functional", "Functionality")

    # ----------------------------------------------------
    # TAB 3: VULNERABILITY
    # ----------------------------------------------------
    create_categorized_sheet("Vulnerability", "Security / Vulnerability")

    # ----------------------------------------------------
    # TAB 4: API UNIT
    # ----------------------------------------------------
    create_categorized_sheet("API Unit", "API Unit")

    # ----------------------------------------------------
    # TAB 5: UI UX
    # ----------------------------------------------------
    create_categorized_sheet("UI UX", "UI UX")

    # ----------------------------------------------------
    # TAB 6: ALL RESULTS
    # ----------------------------------------------------
    create_categorized_sheet("All Results", None)

    # Auto-adjust column widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if cell.row in [1, 2] and ws.title == "Executive Summary":
                    continue
                if len(val) > max_len:
                    # Clip length slightly for extremely long debug strings in column 7
                    max_len = min(len(val), 80)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(excel_path)
    print(f"Excel report generated successfully at: {excel_path}")

if __name__ == "__main__":
    generate_excel()
